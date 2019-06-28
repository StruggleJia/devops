#! /usr/bin/env python
# -*- coding: utf-8 -*-
# code by struggle
# 2019-06-28

import json
from openpyxl import load_workbook



COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333', 'System Foreground', 'System Background' #60-64
)

class ExcelToJson(object):
    def __init__(self, filename):
        self.filename = filename
        self.workfile = load_workbook(self.filename)
        self.sheet = None
        self.cellstatus = {}
        self.allinfo = {}
    
    def get_sheet_name(self):
        return self.workfile.get_sheet_names()

    def select_sheet(self, sheetname):
        self.sheet = self.workfile.get_sheet_by_name(sheetname)

    def sheet_size(self):
        if self.sheet:
            return "数据表最大的坐标\n" + "最大数据的行数:" + str(self.sheet.max_row) + "\n" + "最大数据的列数:" + str(self.sheet.max_column)

    def bgcolor_no(self, start_wz, end_wz):
        if self.sheet:
            for cell in self.sheet[start_wz:end_wz]:
                for i in cell:
                    self.cellstatus[i.fill.start_color.index] = i.value

    def select_sheet_info(self, action="LR"):
        if action == "LR":
            for row in self.sheet.rows:
                for cell in row:
                    info = {}
                    if cell.comment:
                        zb =  str(cell.column)+ "-" + str(cell.row)
                        try:
                            status = self.cellstatus[cell.fill.start_color.index]
                        except:
                            status = "ERROR"
                        rack = cell.value
                        v = cell.comment.text
                        info['status'] = status
                        info['rack'] = rack
                        info['comment'] = v.replace("\n", " ")
                        self.allinfo[zb] = info
            return self.allinfo

if __name__ == "__main__":
    a = ExcelToJson('1111.xlsx')   #excel文件名
    a.select_sheet("2222")         #标签名称
    a.bgcolor_no("M29", "M39")     #图形规则及颜色对应的status_id
    c = a.select_sheet_info()      #处理只有标注的cell信息，生成字典
    print(c)